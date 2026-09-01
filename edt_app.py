# =============================================================================
# IMPORTS UNIFIES
# =============================================================================
import streamlit as st
import pandas as pd
import base64
import io
import time
import re
import os
import hashlib
import zipfile
import math
import smtplib
import secrets
import mimetypes
from datetime import datetime, timedelta, timezone
from collections import defaultdict, Counter
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

try:
    from supabase import create_client, Client
except ImportError:
    create_client = None
    Client = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

try:
    from docx import Document
    from docx.shared import Inches, Pt
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from docx.oxml import OxmlElement
    from docx.oxml.ns import qn
except ImportError:
    Document = None

# ═══════════════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES : Demande EDT + Email Admin
# ═══════════════════════════════════════════════════════════════════════════

def envoyer_email_notification_admin(demande_info, fichier_bytes=None):
    """
    Notifie l'admin par email qu'une nouvelle demande EDT est arrivée.
    """
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email import encoders

        SMTP_SERVER = 'smtp.gmail.com'
        SMTP_PORT = 587
        SMTP_USER = "chef.department.elt.fge@gmail.com"
        SMTP_PASS = "gkzs pdza yodb icvd"
        EMAIL_ADMIN = "chef.department.elt.fge@gmail.com"

        nom_ens = demande_info.get('enseignant_nom', 'Inconnu')
        email_ens = demande_info.get('enseignant_email', 'N/A')
        date_dem = demande_info.get('date_demande', datetime.now().strftime("%d/%m/%Y %H:%M"))
        nb_lignes = len(demande_info.get('lignes', []))

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"📬 Nouvelle demande EDT — {nom_ens}"
        msg["From"] = SMTP_USER
        msg["To"] = EMAIL_ADMIN

        html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:'Segoe UI',Arial,sans-serif;background:#f1f5f9;margin:0;padding:20px;">
<div style="max-width:600px;margin:auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.08);">
<div style="background:linear-gradient(135deg,#DC2626 0%,#EF4444 100%);color:white;padding:25px;text-align:center;">
<h2 style="margin:0;font-size:20px;">🔔 Plateforme EDT — Nouvelle Demande</h2>
<p style="margin:8px 0 0 0;opacity:0.9;font-size:13px;">département d'Électrotechnique - FGE/UDL-SBA</p>
</div>
<div style="padding:30px;">
<p style="color:#334155;font-size:15px;">Salem Admin,</p>
<p style="color:#64748b;font-size:14px;">
    L'enseignant <strong style="color:#1E3A8A;">{nom_ens}</strong> vient de soumettre 
    une demande de mise à jour de son emploi du temps.
</p>
<div style="background:#fef2f2;border-left:5px solid #DC2626;padding:15px;margin:20px 0;border-radius:0 8px 8px 0;">
    <p style="margin:0 0 8px 0;color:#991b1b;font-weight:600;">📋 Récapitulatif</p>
    <table style="width:100%;font-size:13px;color:#334155;">
        <tr><td style="padding:4px 0;"><b>Enseignant :</b></td><td>{nom_ens}</td></tr>
        <tr><td style="padding:4px 0;"><b>Email :</b></td><td>{email_ens}</td></tr>
        <tr><td style="padding:4px 0;"><b>Date :</b></td><td>{date_dem}</td></tr>
        <tr><td style="padding:4px 0;"><b>Créneaux proposés :</b></td><td><strong style="color:#DC2626;">{nb_lignes}</strong> ligne(s)</td></tr>
    </table>
</div>
<div style="background:#eff6ff;border:1px solid #3b82f6;border-radius:8px;padding:15px;margin:20px 0;color:#1e40af;font-size:13px;">
    <strong>⚡ Action requise :</strong><br>
    Connectez-vous à la plateforme, rubrique <strong>"📝 Demandes de Mise à Jour EDT"</strong>.
</div>
<p style="color:#64748b;font-size:13px;"><i>Cet email est généré automatiquement. Le fichier Excel est joint.</i></p>
</div>
<div style="text-align:center;padding:20px;background:#f8fafc;font-size:12px;color:#94a3b8;">
    Faculté de Génie Electrique - Université Djillali Liabes - Sidi Bel Abbes
</div>
</div>
</body></html>"""

        msg.attach(MIMEText(html_body, "html"))

        if fichier_bytes:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(fichier_bytes)
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename=Demande_EDT_{nom_ens.replace(' ', '_')}.xlsx"
            )
            msg.attach(part)

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
        server.quit()
        return True

    except Exception as e:
        print(f"[EMAIL ADMIN] Erreur : {e}")
        return False


def generer_excel_demande_edt(donnees_lignes, nom_enseignant=""):
    """
    Génère un fichier Excel à partir des données de demande.
    Retourne un BytesIO.
    """
    try:
        df = pd.DataFrame(donnees_lignes)
        ordre_cols = ['Enseignements', 'Code', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion']
        cols_presentes = [c for c in ordre_cols if c in df.columns]
        df = df[cols_presentes]

        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Demande EDT', index=False)
        excel_buffer.seek(0)
        return excel_buffer

    except Exception as e:
        st.error(f"Erreur génération Excel : {e}")
        return None


def sauvegarder_demande_edt(email_prof, nom_prof, donnees_lignes, supabase_client):

    """
    Sauvegarde une demande EDT. Si Supabase échoue, bascule en local.
    Garantit l'ordre des colonnes : Enseignements, Code, Enseignants, Horaire, Jours, Lieu, Promotion
    """
    try:
        # Utilise la fonction pro pour garantir l'ordre exact des colonnes
        excel_buffer = generer_excel_demande_edt(donnees_lignes, nom_prof)
        if excel_buffer is None:
            return False, "❌ Erreur lors de la génération Excel", None
        
        excel_buffer.seek(0)
        fichier_bytes = excel_buffer.getvalue()

        date_now = datetime.now()
        fichier_data = {
            "enseignant_email": email_prof,
            "enseignant_nom": nom_prof,
            "lignes": donnees_lignes,
            "date_generation": date_now.strftime("%d/%m/%Y %H:%M")
        }

        supabase_ok = False
        if supabase_client:
            try:
                supabase_client.table("edt_update_requests").insert({
                    "enseignant_id": email_prof,
                    "enseignant_email": email_prof,
                    "enseignant_nom": nom_prof,
                    "fichier_data": fichier_data,
                    "statut": "En attente",
                    "date_demande": date_now.isoformat()
                }).execute()
                supabase_ok = True
            except Exception as e:
                print(f"[SUPABASE] Erreur : {e}")

        if not supabase_ok:
            if "demandes_edt_local" not in st.session_state:
                st.session_state.demandes_edt_local = []
            st.session_state.demandes_edt_local.append({
                "id": len(st.session_state.demandes_edt_local) + 1,
                "enseignant_email": email_prof,
                "enseignant_nom": nom_prof,
                "fichier_data": fichier_data,
                "statut": "En attente",
                "date_demande": date_now.strftime("%d/%m/%Y %H:%M"),
                "excel_bytes": fichier_bytes
            })

        demande_info = {
            "enseignant_nom": nom_prof,
            "enseignant_email": email_prof,
            "date_demande": date_now.strftime("%d/%m/%Y %H:%M"),
            "lignes": donnees_lignes
        }
        email_envoye = envoyer_email_notification_admin(demande_info, fichier_bytes)

        if supabase_ok:
            msg = "✅ Demande enregistrée en ligne."
        else:
            msg = "✅ Demande enregistrée (mode local)."

        if email_envoye:
            msg += " 📧 Admin notifié."
        else:
            msg += " (Email non transmis.)"

        return True, msg, fichier_bytes

    except Exception as e:
        return False, f"❌ Erreur : {str(e)[:150]}", None
        

def generer_excel_demande_edt(donnees_lignes, nom_enseignant=""):
    """
    Génère un fichier Excel à partir des données de demande.
    Retourne un BytesIO prêt pour st.download_button
    """
    import io, pandas as pd
    from datetime import datetime
    
    try:
        df = pd.DataFrame(donnees_lignes)
        
        # Réordonner les colonnes pour un affichage pro
        ordre_cols = ['Enseignements', 'Code', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion']
        cols_presentes = [c for c in ordre_cols if c in df.columns]
        df = df[cols_presentes]
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Demande EDT', index=False)
            ws = writer.sheets['Demande EDT']
            
            # Mise en forme professionnelle
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Ajustement auto des largeurs
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 4, 40)
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        st.error(f"Erreur génération Excel : {e}")
        return None
# =============================================================================
# CONFIGURATION STREAMLIT (UNIQUE)
# =============================================================================
st.set_page_config(
    page_title="Plateforme ELT - UDL-SBA",
    layout="wide",
    page_icon="🏛️",
    initial_sidebar_state="expanded"
)

# Masquer les éléments du menu supérieur
hide_st_style = """
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
.stAppDeployButton {display:none;}
#stDecoration {display:none;}
</style>
"""
st.markdown(hide_st_style, unsafe_allow_html=True)

# =============================================================================
# CONNEXION SUPABASE GLOBALE (partagée)
# =============================================================================
MODE_SUPABASE = False
supabase = None

if create_client:
    try:
        SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
        SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")
        if SUPABASE_URL and SUPABASE_KEY:
            supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
            MODE_SUPABASE = True
    except Exception:
        pass

# =============================================================================
# CONSTANTES COMMUNES
# =============================================================================
_BASE_DIR = Path(__file__).parent.resolve()

FILE_ETUDIANTS = str(_BASE_DIR / "Liste des étudiants_2026-2027.xlsx")
FILE_EDT       = str(_BASE_DIR / "dataEDT-ELT-S1-2027.xlsx")
FILE_ENS       = str(_BASE_DIR / "Permanents-Vacataires-ELT2-2026-2027.xlsx")
NOM_FICHIER_FIXE = FILE_EDT
NOM_FICHIER_CONTACTS = FILE_ENS

HORAIRES_LIST = [
    "8h - 9h30", "9h30 - 11h", "11h - 12h30", "12h30 - 14h", "14h - 15h","14h - 15h30","15h - 16h", "15h30 - 17h"
]
JOURS_SEMAINE = ["Dimanche", "Lundi", "Mardi", "Mercredi", "Jeudi"]

CAUSES_ABSENCES = [
    "Non justifiee",
    "Décès dans l'ascendance, la déscendance ou la parenté",
    "Mariage de l'interessé(e)",
    "Congé de paternité ou de maternité de l'interessé(e)",
    "Mission ou convocation officielle",
    "Maladie de l'interessé(e)",
    "Autres"
]

CODE_ADMIN = "1234"
CODE_ADMIN_EDT = "doctorat2026"

# =============================================================================
# CSS PERSONNALISÉ & STYLES
# =============================================================================
st.markdown("""
<style>
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1e3a8a 0%, #3b82f6 100%);
        min-height: 100vh;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        color: white !important;
    }
    [data-testid="stSidebar"] button {
        width: 100%;
        text-align: left;
    }
    .sidebar-active {
        background: linear-gradient(90deg, #4f46e5, #6366f1) !important;
        color: white !important;
        border-radius: 8px !important;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 15px;
        border-radius: 10px;
        color: white;
    }
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# =============================================================================
# SESSION STATE POUR NAVIGATION INTELLIGENTE
# =============================================================================
if "page_active" not in st.session_state:
    st.session_state.page_active = "accueil_suivi"
if "module_sel" not in st.session_state:
    st.session_state.module_sel = "📊 Suivi d'Assiduite"

# Structure de navigation intelligente
MODULES_NAVIGATION = {
    "📊 Suivi d'Assiduite": {
        "icon": "📊",
        "pages": [
            ("🏠 Accueil", "accueil_suivi"),
            ("📋 Saisir Absences", "saisir_abs"),
            ("📊 Statistiques", "stats"),
            ("📈 Rapports", "rapports"),
            ("👥 Répertoire", "repertoire"),
            ("🔔 Alertes", "alertes")
        ]
    },
    "📅 Gestion des EDTs & Admin": {
        "icon": "📅",
        "pages": [
            ("🏠 Accueil Admin", "accueil_admin"),
            ("📅 Créer EDTs", "creer_edt"),
            ("👨‍🏫 Enseignants", "gerer_ens"),
            ("🎓 Promotions", "gerer_promo"),
            ("⚙️ Paramètres", "parametres")
        ]
    },
    "🧠 EDT Intelligent": {
        "icon": "🧠",
        "pages": [
            ("🏠 Accueil IA", "accueil_ia"),
            ("📊 Affichage EDT", "affichage"),
            ("🔍 Recherche", "recherche"),
            ("📤 Import/Export", "import_export"),
            ("📊 Analytics", "analytics")
        ]
    }
}

# =============================================================================
# SIDEBAR PRINCIPALE - BARRE LATÉRALE INTELLIGENTE
# =============================================================================
with st.sidebar:
    # En-tête
    st.markdown("""
    <div style='text-align: center; padding: 20px 0;'>
        <h1 style='color: white; margin: 0; font-size: 36px;'>🏛️</h1>
        <h2 style='color: white; margin: 10px 0 0 0; font-size: 24px;'>UDL-SBA</h2>
        <p style='color: rgba(255,255,255,0.85); margin: 5px 0 0 0; font-size: 11px;'>
            Département d'Électrotechnique - FGE
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Sélecteur de modules
    st.markdown("""
    <p style='color: white; font-weight: bold; margin-bottom: 10px; margin-top: 10px;'>📂 MODULES</p>
    """, unsafe_allow_html=True)
    
    for module_name, module_info in MODULES_NAVIGATION.items():
        is_active = st.session_state.module_sel == module_name
        
        if st.button(
            f"{module_info['icon']} {module_name}",
            use_container_width=True,
            key=f"module_btn_{module_name.replace(' ', '_')}"
        ):
            st.session_state.module_sel = module_name
            st.session_state.page_active = module_info['pages'][0][1]
            st.rerun()
    
    st.markdown("---")
    
    # Pages du module actif
    st.markdown("""
    <p style='color: white; font-weight: bold; margin-bottom: 10px;'>📄 PAGES</p>
    """, unsafe_allow_html=True)
    
    for module_name, module_info in MODULES_NAVIGATION.items():
        if st.session_state.module_sel == module_name:
            for page_name, page_key in module_info['pages']:
                is_active_page = st.session_state.page_active == page_key
                
                if is_active_page:
                    st.markdown(f"""
                    <div style='
                        background: linear-gradient(90deg, #4f46e5, #6366f1);
                        color: white;
                        padding: 10px 12px;
                        border-radius: 8px;
                        font-weight: 600;
                        margin: 5px 0;
                        text-align: left;
                    '>{page_name}</div>
                    """, unsafe_allow_html=True)
                else:
                    if st.button(
                        page_name,
                        use_container_width=True,
                        key=f"page_btn_{page_key}"
                    ):
                        st.session_state.page_active = page_key
                        st.rerun()
            break
    
    st.markdown("---")
    
    # Infos
    st.markdown("""
    <p style='color: white; font-weight: bold; margin-bottom: 10px;'>ℹ️ INFOS</p>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Année", "2026-2027", label_visibility="collapsed")
    with col2:
        st.metric("Semestre", "S1", label_visibility="collapsed")
    
    st.markdown("---")
    
    # Actions rapides
    st.markdown("""
    <p style='color: white; font-weight: bold; margin-bottom: 10px;'>⚡ ACTIONS</p>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("🔄", use_container_width=True, help="Rafraîchir"):
            st.rerun()
    with col2:
        if st.button("🌓", use_container_width=True, help="Thème"):
            st.info("🌓 Thème switcher - À implémenter")
    with col3:
        if st.button("⚙️", use_container_width=True, help="Paramètres"):
            st.session_state.module_sel = "📅 Gestion des EDTs & Admin"
            st.session_state.page_active = "parametres"
            st.rerun()
    
    st.markdown("---")
    
    # Footer sidebar
    st.markdown("""
    <div style='text-align: center; color: rgba(255,255,255,0.7); font-size: 9px; padding: 15px 0 0 0;'>
        <p style='margin: 0;'>v2.0 | Année 2026-2027</p>
        <p style='margin: 0;'>© Département ELT</p>
        <p style='margin: 5px 0 0 0; font-size: 8px;'>All rights reserved</p>
    </div>
    """, unsafe_allow_html=True)

# Récupérer le module sélectionné
module_sel = st.session_state.module_sel

# =============================================================================
# FONCTIONS UTILITAIRES COMMUNES
# =============================================================================
def nettoyer_nom_enseignant(nom):
    n = str(nom).strip()
    for prefix in ["Pr ", "Dr ", "Mme ", "Mr ", "Dr. ", "Pr. ", "M. "]:
        if n.startswith(prefix):
            n = n[len(prefix):]
    return n.strip()

def extraire_nom_famille(nom_complet):
    n = nettoyer_nom_enseignant(nom_complet)
    parts = n.split()
    if not parts:
        return ""
    return parts[0].upper()

def mapper_promotion(promo_edt):
    p = str(promo_edt).strip().upper()
    # Mapping direct sans doublons ni erreurs
    mapping_direct = {
        "ING1": "ING1",
        "ING2RSE": "ING2",
        "ING3EI": "ING3EI",
        "ING3RSE": "ING3RSE",
        "ING4EI": "ING4",
        "ING4RSE": "ING4RSE",
        "ING5RSE": "ING5RSE",
        "L1MCIL": "L1MCIL",
        "L2ELT": "L2ELT",
        "L2MCIL": "MCIL2",
        "L3ELT": "L3ELT",
        "MCIL2": "MCIL2",
        "MCIL3": "MCIL3",
        "M1CE": "M1CE",
        "M1ER": "M1ER",
        "M1MCIL": "M1MCIL",
        "M1ME": "M1ME",
        "M1RE": "M1RE",
        "M2CE": "M2CE",
        "M2ER": "M2ER",
        "M2MCIL": "M2MCIL",
        "M2ME": "M2ME",
        "M2RE": "M2RE",
    }
    if p in mapping_direct:
        return mapping_direct[p]
    for key, val in mapping_direct.items():
        if key in p or p in key:
            return val
    if "ING1" in p: return "ING1"
    elif "ING2" in p: return "ING2"
    elif "ING3" in p: return "ING3RSE" if "RSE" in p else "ING3EI"
    elif "ING4" in p: return "ING4"
    elif "L1" in p and "MCIL" in p: return "L1MCIL"
    elif "L2" in p and "ELT" in p: return "L2ELT"
    elif "L2" in p and "MCIL" in p: return "MCIL2"
    elif "L3" in p and "ELT" in p: return "L3ELT"
    elif "MCIL3" in p: return "MCIL3"
    elif "M1" in p:
        for code in ["CE", "ER", "MCIL", "ME", "RE"]:
            if code in p: return f"M1{code}"
    elif "M2" in p:
        for code in ["CE", "ER", "MCIL", "ME", "RE"]:
            if code in p: return f"M2{code}"
    return p

def trouver_matiere_promo(nom_ens_complet, df_edt):
    nom_fam = extraire_nom_famille(nom_ens_complet)
    if not nom_fam or df_edt.empty:
        return pd.DataFrame()
    mask = df_edt["Enseignants"].astype(str).str.upper().str.contains(
        re.escape(nom_fam), na=False, regex=True
    )
    df_filtre = df_edt[mask].copy()
    if df_filtre.empty:
        return pd.DataFrame()
    df_filtre["Promotion_Mappee"] = df_filtre["Promotion"].apply(mapper_promotion)
    df_filtre = df_filtre[df_filtre["Enseignants"].astype(str).str.strip().str.lower() != "non defini"]
    return df_filtre

def trouver_toutes_promos_matiere(nom_matiere, df_edt, df_etu):
    """
    ✨ NOUVEAU: Trouve TOUTES les promotions qui ont une matière commune
    Retourne: (promos_list, df_etudiants_tous, promo_primaire)
    """
    if df_edt.empty or nom_matiere.strip() == "":
        return [], pd.DataFrame(), ""
    
    # 1. Trouver TOUTES les lignes de l'EDT avec cette matière
    mask_matiere = df_edt["Enseignements"].astype(str).str.strip().str.lower() == nom_matiere.strip().lower()
    df_mat = df_edt[mask_matiere].copy()
    
    if df_mat.empty:
        return [], pd.DataFrame(), ""
    
    # 2. Extraire TOUTES les promotions uniques pour cette matière
    promos_edt = []
    for promo_brute in df_mat["Promotion"].dropna().unique():
        promo_mapped = mapper_promotion(str(promo_brute).strip())
        promos_edt.append(promo_mapped)
    
    promos_edt = list(set(promos_edt))  # Supprimer les doublons
    
    if not promos_edt:
        return [], pd.DataFrame(), ""
    
    # 3. Chercher les correspondances dans df_etu
    promos_etu_uniques = df_etu["Promotion"].dropna().astype(str).str.strip().unique()
    promos_finales = []
    
    for p_edt in promos_edt:
        p_edt_upper = p_edt.upper()
        # Correspondance exacte
        if p_edt_upper in promos_etu_uniques:
            promos_finales.append(p_edt)
        else:
            # Correspondance partielle
            for p_etu in promos_etu_uniques:
                p_etu_upper = p_etu.upper()
                if p_edt_upper == p_etu_upper or p_edt_upper in p_etu_upper or p_etu_upper in p_edt_upper:
                    promos_finales.append(p_etu)  # Prendre la valeur exacte du fichier
                    break
    
    promos_finales = list(set(promos_finales))  # Supprimer les doublons
    
    if not promos_finales:
        return [], pd.DataFrame(), ""
    
    # 4. Récupérer TOUS les étudiants de CES promotions
    df_etudiants_tous = pd.DataFrame()
    for promo_final in promos_finales:
        df_promo = df_etu[df_etu["Promotion"].astype(str).str.strip().str.upper() == promo_final.upper()]
        df_etudiants_tous = pd.concat([df_etudiants_tous, df_promo], ignore_index=True)
    
    # Supprimer les doublons
    df_etudiants_tous = df_etudiants_tous.drop_duplicates(subset=["Nom_Complet"])
    
    return promos_finales, df_etudiants_tous, promos_finales[0] if promos_finales else ""

def Génerer_page_html(df_data, titre_bilan, colonnes, entetes):
    html_doc = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>{titre_bilan}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
               background-color: #f1f5f9; color: #1e293b; padding: 30px; margin: 0; }}
        .header {{ background-color: #1e3a8a; color: white; padding: 20px;
                   border-radius: 8px 8px 0 0; }}
        .header h1 {{ margin: 0; font-size: 1.5rem; }}
        .content {{ background: white; padding: 20px; border-radius: 0 0 8px 8px;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.05); }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
        th {{ background-color: #0f172a; color: white; padding: 12px 15px; text-align: left; }}
        td {{ padding: 12px 15px; border-bottom: 1px solid #cbd5e1; font-size: 0.9rem; }}
        tr:nth-child(even) {{ background-color: #f8fafc; }}
        .abs-count {{ color: #b91c1c; font-weight: bold; }}
        .badge-fav {{ background-color: #dcfce7; color: #166534; padding: 4px 8px;
                      border-radius: 4px; font-weight: 600; font-size: 0.85em; }}
        .badge-def {{ background-color: #fee2e2; color: #991b1b; padding: 4px 8px;
                      border-radius: 4px; font-weight: 600; font-size: 0.85em; }}
        .badge-att {{ background-color: #fef3c7; color: #92400e; padding: 4px 8px;
                      border-radius: 4px; font-weight: 600; font-size: 0.85em; }}
        .footer {{ text-align: center; margin-top: 25px; font-size: 0.8rem; color: #64748b; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 {titre_bilan}</h1>
        <p>Suivi d'Assiduite - département d'Electrotechnique - UDL-SBA</p>
    </div>
    <div class="content">
        <p>Genere le : {datetime.now().strftime('%d/%m/%Y a %H:%M')}</p>
        <table>
            <thead><tr>"""
    for h in entetes:
        html_doc += f"<th>{h}</th>"
    html_doc += "</tr></thead><tbody>"
    for _, row in df_data.iterrows():
        html_doc += "<tr>"
        for col in colonnes:
            val = row.get(col, "")
            if "Absences" in str(col) or "Total" in str(col):
                html_doc += f"<td class='abs-count'>{val}</td>"
            elif str(col).lower() == "statut":
                if "favor" in str(val).lower() and "de" not in str(val).lower():
                    html_doc += f"<td><span class='badge-fav'>{val}</span></td>"
                elif "defavor" in str(val).lower():
                    html_doc += f"<td><span class='badge-def'>{val}</span></td>"
                else:
                    html_doc += f"<td><span class='badge-att'>{val}</span></td>"
            else:
                html_doc += f"<td>{val}</td>"
        html_doc += "</tr>"
    html_doc += """</tbody></table></div>
    <div class="footer">&copy; 2026 département d'Electrotechnique - UDL-SBA</div>
</body>
</html>"""
    return html_doc

def detecter_colonnes_etudiant(df):
    """Détecte automatiquement les colonnes étudiantes avec tolérance maximale."""
    import unicodedata
    
    def normalize_col(name):
        if pd.isna(name):
            return ""
        s = str(name).strip().lower()
        s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
        s = s.replace(' ', '').replace('-', '').replace('_', '').replace('.', '').replace('/', '')
        return s
    
    cols_norm = {normalize_col(c): c for c in df.columns}
    
    def find_col(variants):
        """Cherche la MEILLEURE correspondance pour éviter les collisions (ex: N° vs Date de naiss.)."""
        best_match = None
        best_score = 0
        MIN_LEN = 4  # Rejette les correspondances sur moins de 4 caractères
        
        for v in variants:
            v_norm = normalize_col(v)
            if not v_norm:
                continue
            
            for key, orig in cols_norm.items():
                if not key:
                    continue
                score = 0
                
                # Correspondance exacte = score maximal
                if v_norm == key:
                    score = 1000 + len(key)
                # Correspondance partielle : exige des chaînes suffisamment longues
                elif len(v_norm) >= MIN_LEN and len(key) >= MIN_LEN:
                    if v_norm in key:
                        score = 500 + len(v_norm)
                    elif key in v_norm:
                        score = 200 + len(key)
                
                if score > best_score:
                    best_score = score
                    best_match = orig
        
        return best_match
    
    mapping = {}
    mapping['nom']           = find_col(['nom', 'name', 'familyname'])
    mapping['prenom']        = find_col(['prenom', 'firstname', 'givenname'])
    mapping['email']         = find_col(['email', 'e-mail', 'mail', 'courriel', 'adressemail'])
    mapping['promotion']     = find_col(['promotion', 'promo', 'niveau', 'annee'])
    mapping['mat_bac']       = find_col(['matbac', 'matriculebac', 'nombac', 'numbac', 'matriculedebac', 'mat.bac'])
    mapping['mat_etud']      = find_col(['matetudiant', 'matriculeetudiant', 'numetudiant', 'netudiant', 'matetud', 'nometudiant', 'codeetudiant'])
    mapping['groupe']        = find_col(['groupe', 'grp', 'group', 'section'])
    mapping['sous_groupe']   = find_col(['sousgroupe', 'sousgrp', 'sg', 'subgroup', 'sousgroupe', 'sousgroupe'])
    mapping['date_naiss']    = find_col(['datedenaissance', 'datenaiss', 'datenaissance', 'naissance', 'datenaiss.', 'datedenaiss.', 'birthdate', 'birth', 'daten'])
    mapping['lieu_naiss']    = find_col(['lieudenaissance', 'lieunaiss', 'lieunaissance', 'lieunaiss.', 'lieudenaiss.', 'birthplace', 'lieu'])
    # ✨ NOUVELLES COLONNES
    mapping['admis_dette']   = find_col(['admisdette', 'admis_dette', 'admisdette', 'endette', 'en_dette', 'dette'])
    mapping['conge_acad']    = find_col(['congeacademique', 'conge_academique', 'congeacad', 'conge_acad', 'congee', 'conge'])
    
    return mapping


def format_date_naissance(val):
    """Formate une date de naissance quelle que soit son type d'origine (Excel, string, Timestamp...)."""
    if pd.isna(val):
        return 'N/A'
    
    # Cas 1 : Déjà un objet date/datetime/pandas Timestamp
    if hasattr(val, 'strftime'):
        try:
            return val.strftime('%d/%m/%Y')
        except:
            pass
    
    # Cas 2 : Nombre → Date Excel (nombre de jours depuis 1899-12-30)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            dt = pd.to_datetime(val, unit='D', origin='1899-12-30')
            return dt.strftime('%d/%m/%Y')
        except:
            pass
    
    # Cas 3 : Chaîne de caractères
    val_str = str(val).strip()
    if val_str:
        import re
        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', val_str):
            try:
                dt = pd.to_datetime(val_str, dayfirst=True, format='%d/%m/%Y')
                return dt.strftime('%d/%m/%Y')
            except:
                pass
        try:
            dt = pd.to_datetime(val_str, dayfirst=True, errors='raise')
            return dt.strftime('%d/%m/%Y')
        except:
            pass
    
    return str(val)
# =============================================================================
# MODULE 1 : SUIVI Assiduité DES ETUDIANTS
# =============================================================================
# FONCTION DE LECTURE EXCEL ROBUSTE
    # =============================================================================
def lire_excel_robuste(chemin_ou_fichier, sheet_name=0):
    """Lit un fichier Excel en essayant plusieurs engines (.xlsx, .xls, .xlsb)."""
    if chemin_ou_fichier is None:
        return None
    
    if hasattr(chemin_ou_fichier, 'seek'):
        chemin_ou_fichier.seek(0)
    
    # Détection prioritaire selon l'extension
    nom = ""
    if hasattr(chemin_ou_fichier, 'name'):
        nom = chemin_ou_fichier.name.lower()
    elif isinstance(chemin_ou_fichier, str):
        nom = os.path.basename(chemin_ou_fichier).lower()
    
    # Ordre des engines : xlrd prioritaire pour les .xls anciens
    engines = ['openpyxl', 'xlrd', 'pyxlsb']
    if nom.endswith('.xls') and not nom.endswith('.xlsx'):
        engines = ['xlrd', 'openpyxl', 'pyxlsb']
    
    last_err = None
    for engine in engines:
        try:
            if hasattr(chemin_ou_fichier, 'seek'):
                chemin_ou_fichier.seek(0)
            return pd.read_excel(chemin_ou_fichier, sheet_name=sheet_name, engine=engine)
        except Exception as e:
            last_err = e
            continue
            
    raise ValueError(f"❌ Format non reconnu. Utilisez un fichier Excel valide (.xlsx, .xls, .xlsb). Erreur : {last_err}")


def run_Assiduité():
    import io
    from fpdf import FPDF
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    st.title("📊 Plateforme de gestion des emplois du temps & Suivi d'Assiduité des Étudiants")
    st.caption("département d'Electrotechnique - Faculté de génie Electrique - UDL-SBA - année 2026-2027")
    
    # =============================================================================
    # CHARGEMENT DES DONNÉES (UNIFIÉ)
    # =============================================================================
    fichiers_locaux_ok = all(os.path.exists(c) for c in [FILE_ETUDIANTS, FILE_EDT, FILE_ENS])
    
    df_etu, df_edt, df_ens = pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    if not fichiers_locaux_ok:
        st.warning("⚠️ Fichiers locaux manquants. Veuillez uploader les 3 fichiers Excel :")
        c1, c2, c3 = st.columns(3)
        with c1:
            up_etu = st.file_uploader("Liste des étudiants", type=["xlsx", "xls", "xlsb"], key="up_etu")
        with c2:
            up_edt = st.file_uploader("Données EDT", type=["xlsx", "xls", "xlsb"], key="up_edt")
        with c3:
            up_ens = st.file_uploader("Liste enseignants", type=["xlsx", "xls", "xlsb"], key="up_ens")
        
        if not all([up_etu, up_edt, up_ens]):
            st.info("📤 En attente des fichiers...")
            return
        
        try:
            df_etu = lire_excel_robuste(up_etu)
            df_etu.columns = df_etu.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de lecture étudiants : {e}")
            return
        try:
            df_edt = lire_excel_robuste(up_edt)
            df_edt.columns = df_edt.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de lecture EDT : {e}")
            return
        try:
            df_ens = lire_excel_robuste(up_ens, sheet_name=0)
            df_ens.columns = df_ens.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de lecture enseignants : {e}")
            return
    else:
        try:
            df_etu = lire_excel_robuste(FILE_ETUDIANTS)
            df_etu.columns = df_etu.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de chargement étudiants : {e}")
            return
        try:
            df_edt = lire_excel_robuste(FILE_EDT)
            df_edt.columns = df_edt.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de chargement EDT : {e}")
            return
        try:
            df_ens = lire_excel_robuste(FILE_ENS, sheet_name=0)
            df_ens.columns = df_ens.columns.str.strip()
        except Exception as e:
            st.error(f"❌ Erreur de chargement enseignants : {e}")
            return

    # Vérification finale
    if df_etu.empty or df_edt.empty or df_ens.empty:
        st.error("❌ Données incomplètes après chargement. Vérifiez vos fichiers source.")
        return
    
    # ... suite de votre code ...    
    
    # --- Preparation des listes ---
    if "NOM" in df_ens.columns and "PRÉNOM" in df_ens.columns:
        df_ens["Nom_Complet"] = df_ens["NOM"].astype(str).str.strip().str.upper() + " " + df_ens["PRÉNOM"].astype(str).str.strip().str.title()
        LISTE_PROFS = sorted(df_ens["Nom_Complet"].dropna().unique().tolist())
    elif "Nom" in df_ens.columns and "Prénom" in df_ens.columns:
        df_ens["Nom_Complet"] = df_ens["Nom"].astype(str).str.strip().str.upper() + " " + df_ens["Prénom"].astype(str).str.strip().str.title()
        LISTE_PROFS = sorted(df_ens["Nom_Complet"].dropna().unique().tolist())
    else:
        LISTE_PROFS = []

    # Détection automatique des colonnes Nom/Prénom (insensible à la casse)
    col_nom = None
    col_prenom = None
    for c in df_etu.columns:
        if c.strip().upper() == "NOM":
            col_nom = c
        if c.strip().upper() in ["PRÉNOM", "PRENOM"]:
            col_prenom = c

    if col_nom and col_prenom:
        df_etu["Nom_Complet"] = df_etu[col_nom].astype(str).str.strip().str.upper() + " " + df_etu[col_prenom].astype(str).str.strip().str.title()
    else:
        st.error(f"❌ Colonnes 'Nom' et 'Prénom' introuvables dans le fichier étudiants. Colonnes trouvées : {list(df_etu.columns)}")
        return

    # =============================================================================
    # INITIALISATION SESSION STATE
    # =============================================================================
    if "absences" not in st.session_state:
        st.session_state.absences = []
    if "requetes" not in st.session_state:
        st.session_state.requetes = []
    if "confirm_reset" not in st.session_state:
        st.session_state.confirm_reset = False
    if "confirm_reset_abs" not in st.session_state:
        st.session_state.confirm_reset_abs = False

    # =============================================================================
    if 'étudiant_auth' not in st.session_state:
        st.session_state.étudiant_auth = None
    if 'étudiant_otp' not in st.session_state:
        st.session_state.étudiant_otp = None
    if 'étudiant_otp_email' not in st.session_state:
        st.session_state.étudiant_otp_email = None

    # FONCTIONS SUPABASE
    # =============================================================================
    def charger_absences_supabase(matiere=None, promotion=None):
        if not MODE_SUPABASE:
            return []
        try:
            query = supabase.table("suivi_assiduite_2026").select("*")
            if matiere:
                query = query.eq("matiere", matiere)
            if promotion:
                query = query.eq("promotion", promotion)
            res = query.execute()
            return res.data if res.data else []
        except Exception as e:
            st.error(f"Erreur Supabase (charger absences) : {e}")
            return []

    def enregistrer_absence_supabase(payload):
        if not MODE_SUPABASE:
            return False
        try:
            supabase.table("suivi_assiduite_2026").insert(payload).execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (enregistrer) : {e}")
            return False

    def supprimer_absences_supabase(matiere, promotion):
        if not MODE_SUPABASE:
            return False
        try:
            supabase.table("suivi_assiduite_2026").delete().eq("matiere", matiere).eq("promotion", promotion).execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (supprimer) : {e}")
            return False

    def rehabiliter_absences_etudiant_supabase(etudiant, matiere, date_abs=None, jour_abs=None, horaire_abs=None):
        if not MODE_SUPABASE:
            return False
        try:
            query = supabase.table("suivi_assiduite_2026").update({"justifie": True})\
                .eq("etud_non_eligible", etudiant).eq("matiere", matiere)
            if date_abs: query = query.eq("date_absence", date_abs)
            if jour_abs: query = query.eq("jour_absence", jour_abs)
            if horaire_abs: query = query.eq("horaire_absence", horaire_abs)
            query.execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (rehabilitation) : {e}")
            return False

    def charger_requetes_supabase(statut=None, promotion=None):
        if not MODE_SUPABASE:
            return []
        try:
            query = supabase.table("requetes_absences").select("*")
            if statut:
                query = query.eq("statut", statut)
            if promotion:
                query = query.eq("promotion", promotion)
            res = query.execute()
            return res.data if res.data else []
        except Exception as e:
            st.error(f"Erreur Supabase (charger requetes) : {e}")
            return []

    def enregistrer_requete_supabase(payload):
        if not MODE_SUPABASE:
            return False
        try:
            supabase.table("requetes_absences").insert(payload).execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (requete) : {e}")
            return False

    def mettre_a_jour_statut_requete_supabase(req_id, statut):
        if not MODE_SUPABASE:
            return False
        try:
            supabase.table("requetes_absences").update({"statut": statut}).eq("id", req_id).execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (maj statut) : {e}")
            return False

    def reinitialiser_requetes_supabase():
        if not MODE_SUPABASE:
            return False
        try:
            supabase.table("requetes_absences").delete().neq("id", -1).execute()
            return True
        except Exception as e:
            st.error(f"Erreur Supabase (reset) : {e}")
            return False

    
    def get_absences_étudiant(nom_etudiant):
        if MODE_SUPABASE:
            try:
                res = supabase.table("suivi_assiduite_2026")\
                    .select("id,etud_non_eligible,matiere,promotion,date_absence,jour_absence,horaire_absence,cause_non_eligibilite,justifie")\
                    .eq("etud_non_eligible", nom_etudiant)\
                    .order("id", desc=True)\
                    .limit(50)\
                    .execute()
                return res.data if res.data else []
            except Exception as e:
                st.error(f"Erreur de chargement absences : {e}")
                return []
        else:
            return [a for a in st.session_state.absences if a.get("etud_non_eligible") == nom_etudiant]    
        
    
    def trouver_requete_existante(nom_etudiant, matiere, date_abs=None, jour_abs=None, horaire_abs=None):
        """Retourne une requete EN ATTENTE existante pour empecher les doublons de depot."""
        if MODE_SUPABASE:
            try:
                query = supabase.table("requetes_absences").select("id,statut")\
                    .eq("nom_etudiant", nom_etudiant)\
                    .eq("matiere", matiere)\
                    .eq("statut", "En attente")
                if date_abs: query = query.eq("date_absence", date_abs)
                if jour_abs: query = query.eq("jour_absence", jour_abs)
                if horaire_abs: query = query.eq("horaire_absence", horaire_abs)
                res = query.limit(1).execute()
                return res.data[0] if res.data else None
            except Exception as e:
                st.warning(f"⚠️ Vérification justificatif lente, réessayez : {e}")
                return None
    for r in st.session_state.requetes:
        if (r.get("nom_etudiant") == nom_etudiant and 
            r.get("matiere") == matiere and 
            r.get("statut") == "En attente"):
            if date_abs and r.get("date_absence") != date_abs: continue
            if jour_abs and r.get("jour_absence") != jour_abs: continue
            if horaire_abs and r.get("horaire_absence") != horaire_abs: continue
            return r
        return None        
    for r in st.session_state.requetes:
        if (r.get("nom_etudiant") == nom_etudiant and 
                r.get("matiere") == matiere and 
                r.get("statut") == "En attente"):
                if date_abs and r.get("date_absence") != date_abs: continue
                if jour_abs and r.get("jour_absence") != jour_abs: continue
                if horaire_abs and r.get("horaire_absence") != horaire_abs: continue
                return r    
        return None

    
    def trouver_derniere_requete(nom_etudiant, matiere, date_abs=None, jour_abs=None, horaire_abs=None):
        """Retourne la DERNIERE requete (tous statuts) pour affichage du statut final."""
        if MODE_SUPABASE:
            try:
                query = supabase.table("requetes_absences").select("id,statut,motif,date_demande")\
                    .eq("nom_etudiant", nom_etudiant)\
                    .eq("matiere", matiere)
                if date_abs: query = query.eq("date_absence", date_abs)
                if jour_abs: query = query.eq("jour_absence", jour_abs)
                if horaire_abs: query = query.eq("horaire_absence", horaire_abs)
                res = query.order("id", desc=True).limit(1).execute()
                return res.data[0] if res.data else None
            except Exception as e:
                st.warning(f"⚠️ Chargement statut justificatif lent : {e}")
                return None
        candidates = [r for r in st.session_state.requetes
                      if r.get("nom_etudiant") == nom_etudiant 
                      and r.get("matiere") == matiere]
        filtered = []
        for r in candidates:
            if date_abs and r.get("date_absence") != date_abs: continue
            if jour_abs and r.get("jour_absence") != jour_abs: continue
            if horaire_abs and r.get("horaire_absence") != horaire_abs: continue
            filtered.append(r)
        return filtered[-1] if filtered else None
    def supprimer_derniere_absence_supabase(etudiant, matiere, promotion):
        if not MODE_SUPABASE:
            return False
        try:
            res = supabase.table("suivi_assiduite_2026").select("*")\
                .eq("etud_non_eligible", etudiant)\
                .eq("matiere", matiere)\
                .eq("promotion", promotion)\
                .order("id", desc=True).limit(1).execute()
            if res.data:
                last_id = res.data[0]["id"]
                supabase.table("suivi_assiduite_2026").delete().eq("id", last_id).execute()
                return True
            return False
        except Exception as e:
            st.error(f"Erreur Supabase (annulation) : {e}")
            return False

    def supprimer_derniere_absence_locale(etudiant, matiere, promotion):
        candidates = [
            (idx, a) for idx, a in enumerate(st.session_state.absences)
            if a.get("etud_non_eligible") == etudiant
            and a.get("matiere") == matiere
            and a.get("promotion") == promotion
        ]
        if candidates:
            last_idx = candidates[-1][0]
            st.session_state.absences.pop(last_idx)
            return True
        return False


    # =============================================================================

    def envoyer_otp_étudiant(email_dest, nom_etud, code_otp):
        try:
            import smtplib
            from email.mime.text import MIMEText
            body = f"Bonjour {nom_etud},\n\nVotre code d'accès à la Plateforme de Suivi d'Assiduité est : {code_otp}\n\nCe code est valable 10 minutes.\n\ndépartement d'Électrotechnique - FGE/UDL-SBA"
            msg = MIMEText(body)
            msg["Subject"] = "Code d'accès - Plateforme Assiduité"
            msg["From"] = "chef.department.elt.fge@gmail.com"
            msg["To"] = str(email_dest).strip()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login("chef.department.elt.fge@gmail.com", "gkzs pdza yodb icvd")
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            st.error(f"Erreur envoi email : {e}")
            return False

    def envoyer_notification_absence_étudiant(email_dest, nom_etud, matiere, enseignant, jour, horaire, date_abs, cause, absences_étudiant=None):
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            # ═══ RÉCAPITULATIF DES ABSENCES PAR matiere ═══
            recap_html = ""
            if absences_étudiant:
                matieres_abs = [a.get("matiere", "Inconnue") for a in absences_étudiant]
                compteur = Counter(matieres_abs)
                if compteur:
                    recap_rows = ""
                    for mat, count in sorted(compteur.items()):
                        couleur = "#b91c1c" if count >= 5 else "#d97706" if count >= 3 else "#1e293b"
                        alerte = " 🚨 EXCLU" if count >= 5 else (" ⚠️ Attention" if count >= 3 else "")
                        recap_rows += f"<tr><td style='padding:8px;border-bottom:1px solid #e2e8f0;'>{mat}</td><td style='padding:8px;border-bottom:1px solid #e2e8f0;text-align:center;font-weight:bold;color:{couleur};'>{count}{alerte}</td></tr>"

                    recap_html = f"""
                    <div style="background:#fff7ed;border:1px solid #fdba74;border-radius:8px;padding:15px;margin:20px 0;">
                        <h3 style="margin:0 0 10px 0;color:#9a3412;font-size:15px;">📊 Récapitulatif de vos absences</h3>
                        <table style="width:100%;border-collapse:collapse;font-size:13px;">
                            <thead><tr style="background:#ffedd5;">
                                <th style="padding:8px;text-align:left;border-bottom:2px solid #fdba74;">matiere</th>
                                <th style="padding:8px;text-align:center;border-bottom:2px solid #fdba74;">Nombre d'absences</th>
                            </tr></thead>
                            <tbody>{recap_rows}</tbody>
                        </table>
                    </div>
                    """

            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"Signalement d'absence - {matiere}"
            msg["From"] = "chef.department.elt.fge@gmail.com"
            msg["To"] = str(email_dest).strip()

            html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:#f1f5f9;margin:0;padding:20px;">
<div style="max-width:600px;margin:auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.08);">
<div style="background:linear-gradient(135deg,#1E3A8A 0%,#3B82F6 100%);color:white;padding:25px;text-align:center;">
<h1 style="margin:0;font-size:20px;">département d'Electrotechnique - FGE/UDL-SBA</h1>
<p style="margin:8px 0 0 0;opacity:0.9;font-size:13px;">Plateforme de Suivi d'Assiduité - année 2026-2027</p>
</div>
<div style="background:#fef2f2;border-left:5px solid #ef4444;padding:15px;margin:20px;color:#991b1b;font-weight:600;">
Vous avez été signalé(e) absent(e) lors d'une séance de cours.
</div>
<div style="padding:20px 30px;">
<p style="color:#334155;margin-bottom:20px;">Salem <strong>{nom_etud}</strong>,</p>
<p style="color:#64748b;font-size:14px;">L'enseignant ci-dessous a enregistré votre absence. Voici les details de la séance :</p>
<table style="width:100%;border-collapse:collapse;margin-top:15px;">
<tr style="border-bottom:1px solid #e2e8f0;"><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">matiere</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{matiere}</td></tr>
<tr style="border-bottom:1px solid #e2e8f0;"><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">Charge de cours</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{enseignant}</td></tr>
<tr style="border-bottom:1px solid #e2e8f0;"><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">Date</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{date_abs}</td></tr>
<tr style="border-bottom:1px solid #e2e8f0;"><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">Jour</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{jour}</td></tr>
<tr style="border-bottom:1px solid #e2e8f0;"><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">Horaire</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{horaire}</td></tr>
<tr><td style="padding:10px 0;color:#64748b;font-size:13px;font-weight:600;">Motif enregistre</td><td style="padding:10px 0;color:#1e293b;font-weight:700;text-align:right;">{cause}</td></tr>
</table>

{recap_html}

<div style="background:#fee2e2;border:1px solid #ef4444;border-radius:8px;padding:15px;margin:20px 0;color:#991b1b;font-size:13px;">
    <strong>⏰ Délais de justification :</strong><br>
    Vous disposez d'un délai de <strong>48 heures</strong> à compter de la date d'absence pour déposer un justificatif via l'onglet <strong>Justificatifs</strong> de la plateforme. Passé ce délai, l'absence sera considérée comme <strong>définitivement non justifiee</strong>.
</div>

<p style="color:#64748b;font-size:13px;margin-top:20px;">Accédez à la plateforme pour déposer votre justificatif : onglet <strong>Justificatifs</strong>.</p>
</div>
<div style="text-align:center;padding:20px;background:#f8fafc;font-size:12px;color:#94a3b8;">
Faculté de génie Electrique - Université Djillali Liabes - Sidi Bel Abbes<br>
Cet email est genere automatiquement - merci de ne pas y repondre.
</div>
</div>
</body>
</html>"""
            msg.attach(MIMEText(html_body, "html"))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login("chef.department.elt.fge@gmail.com", "gkzs pdza yodb icvd")
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            st.warning(f"Notification email non envoyee : {e}")
            return False

    def envoyer_notification_decision_etudiant(email_dest, nom_etud, matiere, statut, motif=""):
        """Envoie une notification à l'étudiant après décision admin (Favorable/Défavorable)."""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            couleur = "#166534" if "Favorable" in statut else "#991b1b"
            icone = "✅" if "Favorable" in statut else "❌"
            titre = "Justificatif ACCEPTÉ" if "Favorable" in statut else "Justificatif REJETÉ"
            msg_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:#f1f5f9;margin:0;padding:20px;">
<div style="max-width:600px;margin:auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.08);">
<div style="background:linear-gradient(135deg,#1E3A8A 0%,#3B82F6 100%);color:white;padding:25px;text-align:center;">
<h1 style="margin:0;font-size:20px;">département d'Electrotechnique - FGE/UDL-SBA</h1>
<p style="margin:8px 0 0 0;opacity:0.9;font-size:13px;">Plateforme de Suivi d'Assiduité - année 2026-2027</p>
</div>
<div style="padding:30px;">
<p style="color:#334155;margin-bottom:20px;">Salem <strong>{nom_etud}</strong>,</p>
<div style="background:{'#dcfce7' if 'Favorable' in statut else '#fee2e2'};border-left:5px solid {couleur};padding:15px;margin:20px 0;color:{couleur};font-weight:600;border-radius:0 8px 8px 0;">
    {icone} <strong>{titre}</strong><br>
    matière concernée : <strong>{matiere}</strong>
</div>
<p style="color:#64748b;font-size:14px;"><strong>Motif du justificatif :</strong> {motif if motif else "Non précisé"}</p>
{"<p style='color:#166534;font-size:14px;font-weight:600;'>🎓 Votre absence est désormais justifiee. Vous conservez votre éligibilité à l'examen et vous avez le droit à un <strong>examen de remplacement (rattrapage)</strong> si nécessaire.</p>" if "Favorable" in statut else "<p style='color:#991b1b;font-size:14px;'>Votre demande de justification a été rejetée. L'absence reste non justifiee et les sanctions réglementaires s'appliquent.</p>"}
</div>
<div style="text-align:center;padding:20px;background:#f8fafc;font-size:12px;color:#94a3b8;">
Faculté de génie Electrique - Université Djillali Liabes - Sidi Bel Abbes<br>
Cet email est généré automatiquement - merci de ne pas y répondre.
</div>
</div>
</body>
</html>"""
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"{titre} - {matiere}"
            msg["From"] = "chef.department.elt.fge@gmail.com"
            msg["To"] = str(email_dest).strip()
            msg.attach(MIMEText(msg_body, "html"))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login("chef.department.elt.fge@gmail.com", "gkzs pdza yodb icvd")
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            st.warning(f"Notification décision étudiant non envoyée : {e}")
            return False

    def envoyer_notification_decision_enseignant(email_dest, nom_ens, nom_etud, matiere, statut, promotion=""):
        """Envoie une notification à l'enseignant après décision admin sur une justification."""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            couleur = "#166534" if "Favorable" in statut else "#991b1b"
            icone = "✅" if "Favorable" in statut else "❌"
            titre = "Justificatif ACCEPTÉ" if "Favorable" in statut else "Justificatif REJETÉ"

            msg_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:#f1f5f9;margin:0;padding:20px;">
<div style="max-width:600px;margin:auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.08);">
<div style="background:linear-gradient(135deg,#1E3A8A 0%,#3B82F6 100%);color:white;padding:25px;text-align:center;">
<h1 style="margin:0;font-size:20px;">département d'Electrotechnique - FGE/UDL-SBA</h1>
<p style="margin:8px 0 0 0;opacity:0.9;font-size:13px;">Plateforme de Suivi d'Assiduité - année 2026-2027</p>
</div>
<div style="padding:30px;">
<p style="color:#334155;margin-bottom:20px;">Salem <strong>{nom_ens}</strong>,</p>
<p style="color:#64748b;font-size:14px;">L'administration a rendu son avis sur une demande de justification d'absence concernant un étudiant de votre enseignement.</p>
<div style="background:{'#dcfce7' if 'Favorable' in statut else '#fee2e2'};border-left:5px solid {couleur};padding:15px;margin:20px 0;color:{couleur};font-weight:600;border-radius:0 8px 8px 0;">
    {icone} <strong>{titre}</strong><br>
    <strong>Étudiant :</strong> {nom_etud}<br>
    <strong>matiere :</strong> {matiere}<br>
    <strong>Promotion :</strong> {promotion}
</div>
{"<div style='background:#eff6ff;border:1px solid #3b82f6;border-radius:8px;padding:15px;margin:20px 0;color:#1e40af;font-size:13px;'><strong>📋 Information importante :</strong><br>Cet étudiant a obtenu un avis <strong>favorable</strong> pour la justification de son absence. Il est désormais réhabilité et <strong>a le droit à un examen de remplacement (rattrapage)</strong> s'il le souhaite, conformément au règlement intérieur.</div>" if "Favorable" in statut else "<div style='background:#fff7ed;border:1px solid #fdba74;border-radius:8px;padding:15px;margin:20px 0;color:#9a3412;font-size:13px;'><strong>📋 Information :</strong><br>La demande de justification a été <strong>rejetée</strong>. L'absence reste enregistrée comme non justifiee et l'étudiant reste sous le seuil d'exclusion si applicable.</div>"}
</div>
<div style="text-align:center;padding:20px;background:#f8fafc;font-size:12px;color:#94a3b8;">
Faculté de génie Electrique - Université Djillali Liabes - Sidi Bel Abbes<br>
Cet email est généré automatiquement - merci de ne pas y répondre.
</div>
</div>
</body>
</html>"""
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"[{titre}] {nom_etud} - {matiere}"
            msg["From"] = "chef.department.elt.fge@gmail.com"
            msg["To"] = str(email_dest).strip()
            msg.attach(MIMEText(msg_body, "html"))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login("chef.department.elt.fge@gmail.com", "gkzs pdza yodb icvd")
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            st.warning(f"Notification décision enseignant non envoyée : {e}")
            return False

    def trouver_email_étudiant(nom_etudiant, df_étudiants):
        # 1. Priorité : base de données Supabase (email saisi lors de connexion étudiant)
        if MODE_SUPABASE:
            try:
                res = supabase.table("étudiants_emails").select("email").eq("nom_complet", str(nom_etudiant).strip()).execute()
                if res.data and len(res.data) > 0:
                    email_db = str(res.data[0]["email"]).strip()
                    if email_db and email_db.lower() not in ["nan", "none", ""]:
                        return email_db
            except Exception:
                pass
        # 2. Fallback : fichier Excel source (colonne Email / Mail / E-mail)
        if df_étudiants is None or df_étudiants.empty:
            return None
        col_email = None
        for c in df_étudiants.columns:
            c_up = str(c).strip().upper()
            if c_up in ["EMAIL", "E-MAIL", "MAIL", "COURRIEL", "ADRESSE EMAIL"]:
                col_email = c
                break
        if not col_email:
            return None
        mask = df_étudiants["Nom_Complet"].astype(str).str.strip().str.upper() == str(nom_etudiant).strip().upper()
        match = df_étudiants[mask]
        if not match.empty:
            val = str(match.iloc[0][col_email]).strip()
            if val and val.lower() not in ["nan", "none", ""]:
                return val
        return None

    # =============================================================================
    # AUTHENTIFICATION ETUDIANT (Mat. BAC + OTP)
    # =============================================================================
    # Récupération connexion enseignant (depuis Module 2 EDT)
    user = st.session_state.get("user_data")
    is_enseignant_connecte = user is not None and user.get("role") != "admin"
    is_admin_edt = user is not None and user.get("role") == "admin"

    étudiant_connecte = st.session_state.get("étudiant_auth")

    if not is_enseignant_connecte and not étudiant_connecte and not is_admin_edt:
        st.markdown("<h3 style='text-align:center;color:#1E3A8A;'>🔐 Portail Étudiant</h3>", unsafe_allow_html=True)
        st.info("Accédez à votre espace pour consulter vos absences et déposer des justificatifs.")

        mat_bac_input = st.text_input("🎓 Numéro de Matricule BAC :", key="mat_bac_auth", placeholder="Ex: 12345678")

        if mat_bac_input:
            mat_bac_clean = str(mat_bac_input).strip().upper().replace(" ", "").replace("-", "")
            df_match = pd.DataFrame()
            col_mat_bac = None
            for c in df_etu.columns:
                c_up = str(c).strip().upper().replace('.', '').replace(' ', '').replace('_', '').replace('-', '')
                if "MAT" in c_up and "BAC" in c_up:
                    col_mat_bac = c
                    break

            if col_mat_bac:
                mask_mat = df_etu[col_mat_bac].astype(str).str.strip().str.upper().str.replace(' ', '').str.replace('-', '') == mat_bac_clean
                df_match = df_etu[mask_mat]
            else:
                for c in df_etu.columns:
                    vals = df_etu[c].astype(str).str.strip().str.upper().str.replace(' ', '').str.replace('-', '')
                    if vals.eq(mat_bac_clean).any():
                        df_match = df_etu[vals == mat_bac_clean]
                        break

            if not df_match.empty:
                etud_nom = str(df_match.iloc[0]['Nom_Complet']).strip()
                etud_promo = str(df_match.iloc[0]['Promotion']).strip() if 'Promotion' in df_match.columns else ''
                st.success(f"✅ Étudiant trouvé : **{etud_nom}** ({etud_promo})")

                email_input = st.text_input("📧 Votre adresse email :", key="email_etud_auth", placeholder="ex: nom@email.com")

                if email_input and "@" in str(email_input):
                    if st.button("📧 Recevoir mon code d'accès", use_container_width=True, key="btn_otp"):
                        import random
                        otp_code = str(random.randint(100000, 999999))
                        st.session_state.étudiant_otp = otp_code
                        st.session_state.étudiant_otp_email = str(email_input).strip()

                        sent = envoyer_otp_étudiant(email_input, etud_nom, otp_code)
                        if sent:
                            st.success(f"✅ Code envoyé à : `{email_input}` — Vérifiez votre boîte mail (et les spams).")
                        else:
                            st.warning(f"⚠️ Impossible d'envoyer l'email. Votre code (mode démo) : `{otp_code}`")

                if st.session_state.get("étudiant_otp"):
                    otp_input = st.text_input("🔑 Saisissez le code reçu par email :", type="password", key="otp_input_auth")
                    if st.button("✅ Valider mon accès", use_container_width=True, key="btn_valider_otp"):
                        if otp_input == st.session_state.get("étudiant_otp"):
                            st.session_state.étudiant_auth = {
                                "mat_bac": mat_bac_clean,
                                "nom": etud_nom,
                                "email": st.session_state.étudiant_otp_email,
                                "promotion": etud_promo
                            }
                            # Stockage persistant de l'email dans Supabase pour notifications futures
                            if MODE_SUPABASE:
                                try:
                                    supabase.table("étudiants_emails").upsert({
                                        "nom_complet": etud_nom,
                                        "mat_bac": mat_bac_clean,
                                        "email": st.session_state.étudiant_otp_email,
                                        "promotion": etud_promo,
                                        "derniere_connexion": datetime.now().isoformat()
                                    }, on_conflict="nom_complet").execute()
                                except Exception:
                                    pass
                            st.session_state.étudiant_otp = None
                            st.session_state.étudiant_otp_email = None
                            st.success(f"🎓 Bienvenue {etud_nom} ! Accès autorisé...")
                            time.sleep(0.8)
                            st.rerun()
                        else:
                            st.error("❌ Code incorrect. Veuillez réessayer.")
            else:
                st.error("❌ Matricule BAC non reconnu dans la base étudiants.")

        return

    # =============================================================================
    # ONGLETS
    # =============================================================================
    # Création des onglets (toujours 4 pour éviter UnboundLocalError)
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Suivi d'Assiduité", "📩 Justificatifs", "📊 Bilans & Exports", "👤 Infos Étudiant"])
    with tab1:
        st.header("📝 Suivi de l'Assiduité et Compteur d'Absences")

        sel_prof = ""
        sel_mat = ""
        promo_c = ""
        df_matiere = pd.DataFrame()

        if is_enseignant_connecte:
            sel_prof = user['nom_officiel']
            st.success(f"👤 Bienvenue **{sel_prof}** — Espace Suivi d'Assiduité")
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**Enseignant :** `{sel_prof}`")
            with c2:
                st.markdown("*Accès direct — Aucun code requis*")
        elif is_admin_edt:
            st.success(f"👤 Mode Administrateur — Accès complet au suivi d'Assiduité")
            c1, c2 = st.columns(2)
            with c1:
                sel_prof = st.selectbox("👤 Sélectionnez l'Enseignant :", [""] + LISTE_PROFS, key="prof_admin_assid")

    with tab4:
        st.header("👤 Espace d'Information Étudiant")
        st.caption("Recherche d'un étudiant par son nom complet ou par son numéro de Matricule BAC")

        # Sélection du mode de recherche
        mode_recherche = st.radio(
            "Méthode de recherche :",
            ["🔍 Par Nom & Prénom", "🎓 Par Matricule BAC"],
            horizontal=True,
            key="mode_recherche_etud"
        )

        etud_trouve = None

        if mode_recherche == "🔍 Par Nom & Prénom":
            liste_tous_etudiants = sorted(df_etu["Nom_Complet"].dropna().unique().tolist())
            etud_sel_nom = st.selectbox(
                "Sélectionnez un étudiant dans la liste :",
                [""] + liste_tous_etudiants,
                key="select_etud_nom_tab4"
            )
            if etud_sel_nom:
                etud_trouve = df_etu[df_etu["Nom_Complet"] == etud_sel_nom].iloc[0]

        else:
            mat_input_tab4 = st.text_input(
                "Saisissez le numéro de Matricule BAC :",
                key="input_mat_bac_tab4",
                placeholder="Ex: 202038012345"
            )
            if mat_input_tab4:
                mat_clean = str(mat_input_tab4).strip().upper().replace(" ", "").replace("-", "")
                col_mat = None
                for c in df_etu.columns:
                    c_up = str(c).strip().upper().replace('.', '').replace(' ', '').replace('_', '').replace('-', '')
                    if "MAT" in c_up and "BAC" in c_up:
                        col_mat = c
                        break

                if col_mat:
                    mask = df_etu[col_mat].astype(str).str.strip().str.upper().str.replace(' ', '').str.replace('-', '') == mat_clean
                    match_df = df_etu[mask]
                else:
                    match_df = pd.DataFrame()
                    for c in df_etu.columns:
                        vals = df_etu[c].astype(str).str.strip().str.upper().str.replace(' ', '').str.replace('-', '')
                        if vals.eq(mat_clean).any():
                            match_df = df_etu[vals == mat_clean]
                            break

                if not match_df.empty:
                    etud_trouve = match_df.iloc[0]
                else:
                    st.error("❌ Aucun étudiant trouvé avec ce numéro de Matricule BAC.")

        # Affichage de la fiche étudiant
        if etud_trouve is not None:
            mapping_cols = detecter_colonnes_etudiant(df_etu)

            nom_aff = etud_trouve.get(mapping_cols['nom'], etud_trouve.get("Nom_Complet", "N/A"))
            prenom_aff = etud_trouve.get(mapping_cols['prenom'], "")
            nom_complet_etud = str(etud_trouve.get("Nom_Complet", f"{nom_aff} {prenom_aff}")).strip()

            st.markdown("---")
            st.subheader(f"🎓 Fiche Étudiant : {nom_complet_etud}")

            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f"**Nom :** `{nom_aff}`")
                st.markdown(f"**Prénom :** `{prenom_aff}`")
                st.markdown(f"**Promotion :** `{etud_trouve.get(mapping_cols['promotion'], 'N/A')}`")
            with c2:
                st.markdown(f"**Matricule BAC :** `{etud_trouve.get(mapping_cols['mat_bac'], 'N/A')}`")
                st.markdown(f"**Matricule Étudiant :** `{etud_trouve.get(mapping_cols['mat_etud'], 'N/A')}`")
                st.markdown(f"**Groupe :** `{etud_trouve.get(mapping_cols['groupe'], 'N/A')}`")
            with c3:
                date_n_brute = etud_trouve.get(mapping_cols['date_naiss'], None)
                st.markdown(f"**Date de Naissance :** `{format_date_naissance(date_n_brute)}`")
                st.markdown(f"**Lieu de Naissance :** `{etud_trouve.get(mapping_cols['lieu_naiss'], 'N/A')}`")

                email_etud_trouve = trouver_email_étudiant(nom_complet_etud, df_etu)
                st.markdown(f"**Email :** `{email_etud_trouve if email_etud_trouve else 'Non renseigné'}`")

            # --- HISTORIQUE DES ABSENCES DE L'ÉTUDIANT ---
            st.markdown("### 📊 Historique des Absences Signalées")

            abs_etud = get_absences_étudiant(nom_complet_etud)

            if abs_etud:
                df_abs_etud = pd.DataFrame(abs_etud)

                # Nettoyage des colonnes à afficher
                cols_to_show = []
                rename_dict = {}

                for col, name in [
                    ("matiere", "Matière"),
                    ("promotion", "Promotion"),
                    ("date_absence", "Date Absence"),
                    ("jour_absence", "Jour"),
                    ("horaire_absence", "Horaire"),
                    ("cause_non_eligibilite", "Motif / Cause"),
                    ("justifie", "Justifiée")
                ]:
                    if col in df_abs_etud.columns:
                        cols_to_show.append(col)
                        rename_dict[col] = name

                df_display = df_abs_etud[cols_to_show].rename(columns=rename_dict)

                if "Justifiée" in df_display.columns:
                    df_display["Justifiée"] = df_display["Justifiée"].apply(lambda x: "✅ Oui" if x else "❌ Non")

                st.dataframe(df_display, use_container_width=True)

                # Compteur récapitulatif par matière
                st.markdown("#### 📈 Récapitulatif par Matière :")
                mat_counts = Counter([a.get("matiere", "Inconnue") for a in abs_etud])

                cols_mats = st.columns(min(len(mat_counts), 4) if mat_counts else 1)
                idx_c = 0
                for mat_k, count_v in mat_counts.items():
                    with cols_mats[idx_c % len(cols_mats)]:
                        st.metric(label=f"Matière: {mat_k}", value=f"{count_v} absence(s)")
                    idx_c += 1

            else:
                st.success("🎉 Aucune absence enregistrée pour cet étudiant.")
